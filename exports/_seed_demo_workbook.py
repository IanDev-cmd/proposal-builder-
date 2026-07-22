"""Create DEMO seed workbook: Ops write-back tabs + empty leads (no demo lead rows)."""
from pathlib import Path

from openpyxl import Workbook

exports = Path(__file__).resolve().parent
wb = Workbook()

# Sheet1 — Get Leads DEMO target; blank = 0 leads (intentional)
ws1 = wb.active
ws1.title = "Sheet1"

notes = wb.create_sheet("Nexus Ops Notes")
notes.append(["Created At", "Mode", "Reference", "Email", "Lead Name", "Tag", "Note"])

quotes = wb.create_sheet("Nexus Ops Quotes")
quotes.append(
    [
        "Updated At",
        "Mode",
        "Reference",
        "Email",
        "Lead Name",
        "Quote Id",
        "Status",
        "Version",
        "Title",
        "Event Type",
        "Event Date",
        "Guests",
        "Repeat Client",
        "Selected Upgrades",
        "Template Id",
        "Selected Inserts",
        "Staff Contact",
        "Base Cost",
        "Contingency",
        "Margin",
        "Margin Amount",
        "Cost To Client",
        "Package Cost",
        "VAT",
        "Upgrade Total",
        "Grand Total",
    ]
)

# Header scaffold only — no lead rows (repaint Get Leads DEMO here later if desired)
enq = wb.create_sheet("Enquiry - Lead Data (2026)")
enq_headers = [
    "Live/Dead/ Blacklisted/Booked",
    "Enquiry Date",
    "Name",
    "Main Contact - Job Role",
    "Company Name  (If Applicable - If so, please link to their LinkedIn company page)",
    "Company Sector  (If Applicable)",
    "Main Contact - Email",
    "Main Contact - Number",
    "",
    "Client Reference Number",
    "Source",
    "Email Account Used",
    "Client Relationship  Representative",
    "",
    "Repeat Enquiry:   Has the company enquired with us before?",
    "Repeat Client:  Has the company booked with us?",
    "",
    "Status",
    "Last Action",
    "Next Action",
    "Next Action Date",
    "Next Action Assigned To",
    "",
    "Best time to call  (if they have specified) - list time and/or days they've mentioned",
    "",
    "Market",
    "Event Type",
    "",
    "Year of Event",
    "Full Event Date",
    "On Hold?",
    "Quarter Period - event is being held",
    "Event Date - Flexible?",
    "Requested Event Times",
    "Group Size",
    "Demographic  (Enter manually)",
    "Budget",
    "",
    "How did they hear about us?  This information will be located in the form they submmited",
    "Do they have a preferred pier?",
    "Do they require our support with travel and accomodation?",
    "What vessel/s are we putting forward?",
]
enq_headers.extend(f"Progress {i}" for i in range(1, 51))
enq.append(enq_headers)

seed_path = exports / "WEOTT-Nexus-TEST-Sheets.xlsx"
wb.save(seed_path)
print("SEED", seed_path)
print("sheets", wb.sheetnames)
